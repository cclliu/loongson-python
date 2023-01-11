# 读取excel文件
# 打开工作簿
from openpyxl import load_workbook

wb = load_workbook(r"word\\1.xlsx", read_only=False)

# 从工作簿中获取表单
# 遍历表单名称
for sheet in wb:
    print(sheet.title)

print(wb.sheetnames)  #

ws = wb.active  # 取出活跃的表单

# 先取出每一行，再取出每一行中的每一个单元格来获取值
for row in ws.values:
    for cell in row:
        print(cell,end=' ')
    print()
